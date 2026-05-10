from pathlib import Path
import openpyxl

XLSX = Path(__file__).resolve().parent.parent / "instructions" / "מקרי בוחן עם נוסחים.xlsx"


def load_use_cases() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c else "" for c in rows[0]]

    def col(name_substr: str) -> int:
        for i, h in enumerate(header):
            if name_substr in h:
                return i
        return -1

    i_rc = col("RC")
    i_desc = col("RC Description")
    i_reason = col("Failure Reason")
    i_script = col("נוסח")  # WhatsApp script column

    cases = []
    for idx, row in enumerate(rows[1:], start=1):
        if not row or i_reason < 0 or not row[i_reason]:
            continue
        rc = row[i_rc] if i_rc >= 0 else ""
        desc = row[i_desc] if i_desc >= 0 else ""
        reason = row[i_reason] if i_reason >= 0 else ""
        script = row[i_script] if i_script >= 0 and row[i_script] else ""
        cases.append({
            "id": str(idx),
            "rc": str(rc).strip(),
            "rc_description": str(desc).strip(),
            "failure_reason": str(reason).strip(),
            "script": str(script).strip(),
        })
    return cases


CASES = load_use_cases()


def get_case(case_id: str) -> dict | None:
    for c in CASES:
        if c["id"] == case_id:
            return c
    return None
